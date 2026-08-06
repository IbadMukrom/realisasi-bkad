"""
Modul untuk mengelola data anggaran — simpan, update, hapus ke file Excel atau Google Sheets.
"""
import pandas as pd
import os
import streamlit as st

from typing import Optional, List, Dict, Any, Tuple

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")
DEFAULT_FILE = os.path.join(DATA_DIR, "dummy_data.xlsx")

NAMA_OPD = "Badan Keuangan dan Aset Daerah (BKAD)"

JENIS_BELANJA_OPTIONS = [
    "Belanja Pegawai",
    "Belanja Barang dan Jasa",
    "Belanja Modal",
    "Belanja Hibah",
    "Belanja Bantuan Sosial",
    "Belanja Tidak Terduga",
    "Belanja Transfer",
]

NAMA_BULAN = {
    1: "Januari", 2: "Februari", 3: "Maret",
    4: "April", 5: "Mei", 6: "Juni",
    7: "Juli", 8: "Agustus", 9: "September",
    10: "Oktober", 11: "November", 12: "Desember",
}


def is_gsheets_configured() -> bool:
    """Memeriksa apakah koneksi Google Sheets sudah dikonfigurasi di secrets."""
    try:
        return "connections" in st.secrets and "gsheets" in st.secrets["connections"]
    except Exception:
        return False


def fix_private_key(pk: str) -> str:
    """
    Membersihkan dan memformat private_key Google Service Account agar kompatibel
    dengan library cryptography dan gspread.
    """
    if not isinstance(pk, str):
        return pk

    pk = pk.strip().strip('"').strip("'")
    pk = pk.replace("\\\\n", "\n").replace("\\n", "\n")

    lines = pk.strip().splitlines()
    cleaned_lines = []
    for line in lines:
        l = line.strip()
        if not l:
            continue
        if l.startswith("-----BEGIN") or l.startswith("-----END"):
            cleaned_lines.append(l)
        else:
            fixed_body = l.replace("_", "/").replace("-", "+")
            cleaned_lines.append(fixed_body)

    return "\n".join(cleaned_lines)


def get_gsheets_client_and_sheet():
    """
    Mendapatkan objek gspread worksheet dari Google Sheets dengan sanitasi private_key otomatis.
    """
    import gspread
    from google.oauth2.service_account import Credentials

    sec = dict(st.secrets["connections"]["gsheets"])
    spreadsheet_url = str(sec.get("spreadsheet", "")).strip()

    creds_dict = {k: v for k, v in sec.items() if k != "spreadsheet"}
    if "private_key" in creds_dict and isinstance(creds_dict["private_key"], str):
        creds_dict["private_key"] = fix_private_key(creds_dict["private_key"])

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]

    creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    gc = gspread.authorize(creds)

    if spreadsheet_url.startswith("http"):
        sh = gc.open_by_url(spreadsheet_url)
    else:
        sh = gc.open(spreadsheet_url)

    try:
        ws = sh.worksheet("Realisasi Anggaran")
    except Exception:
        ws = sh.sheet1
    return ws


def get_all_jenis_belanja(filepath: Optional[str] = None) -> List[str]:
    """
    Mendapatkan daftar seluruh jenis belanja dari Excel/Google Sheets + default options.
    """
    df = load_raw_data(filepath)
    existing = []
    if not df.empty and "jenis_belanja" in df.columns:
        existing = [str(x).strip() for x in df["jenis_belanja"].dropna().unique() if str(x).strip()]
    
    combined = list(dict.fromkeys(JENIS_BELANJA_OPTIONS + existing))
    return sorted(combined)


def get_all_penanggungjawab(filepath: Optional[str] = None) -> List[str]:
    """
    Mendapatkan daftar Bidang Penanggung Jawab dari Excel/Google Sheets.
    """
    defaults = [
        "Sekretariat",
        "Bidang Anggaran",
        "Bidang Perbendaharaan",
        "Bidang Akuntansi & Pelaporan",
        "Bidang Pengelolaan BMD",
    ]
    df = load_raw_data(filepath)
    if not df.empty and "penanggungjawab" in df.columns:
        existing = [str(x).strip() for x in df["penanggungjawab"].dropna().unique() if str(x).strip()]
        combined = list(dict.fromkeys(defaults + existing))
        return sorted(combined)
    return defaults


def get_data_path() -> str:
    """Mendapatkan path file data Excel."""
    return DEFAULT_FILE


@st.cache_data(ttl="10s", show_spinner=False)
def load_raw_data(filepath: Optional[str] = None) -> pd.DataFrame:
    """
    Membaca data mentah dari Google Sheets (jika dikonfigurasi) atau dari file Excel.
    Menggunakan cache 10 detik untuk mencegah error 429 (Quota Exceeded).
    """
    default_cols = ["tahun", "nama_opd", "penanggungjawab", "kode_rekening", "jenis_belanja", "bulan", "pagu_anggaran", "realisasi"]

    if is_gsheets_configured():
        try:
            ws = get_gsheets_client_and_sheet()
            try:
                records = ws.get_all_records()
                if records:
                    df = pd.DataFrame(records)
                else:
                    headers = ws.row_values(1)
                    cols = headers if headers else default_cols
                    df = pd.DataFrame(columns=cols)
            except Exception:
                df = pd.DataFrame(columns=default_cols)

            df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
            return df
        except Exception as e:
            err_name = type(e).__name__
            err_msg = str(e).strip() or repr(e)
            full_err = f"{err_name}: {err_msg}"
            
            if "API has not been used" in full_err or "disabled" in full_err.lower():
                st.warning(f"⚠️ **API Google Belum Aktif di GCP**: Silakan buka Google Cloud Console dan aktifkan (Enable) **Google Sheets API** & **Google Drive API**.\n\nDetail: {full_err}")
            else:
                st.warning(f"⚠️ **Gagal Koneksi Google Sheets**: {full_err}")

    path = filepath or DEFAULT_FILE
    if not os.path.exists(path):
        os.makedirs(os.path.dirname(path), exist_ok=True)
        empty_df = pd.DataFrame(columns=default_cols)
        empty_df.to_excel(path, index=False, sheet_name="Realisasi Anggaran")
        return empty_df

    try:
        from utils.data_loader import read_smart_excel
        df = read_smart_excel(path)
    except Exception:
        try:
            df = pd.read_excel(path)
        except Exception:
            df = pd.DataFrame(columns=default_cols)

    df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")

    if "pagu_anggaran" not in df.columns:
        if "pagu_tahunan" in df.columns:
            df["pagu_anggaran"] = df["pagu_tahunan"]
        elif "pagu" in df.columns:
            df["pagu_anggaran"] = df["pagu"]
        else:
            df["pagu_anggaran"] = 0.0

    if "realisasi" not in df.columns:
        df["realisasi"] = 0.0

    return df


def save_data(df: pd.DataFrame, filepath: Optional[str] = None) -> None:
    """
    Menyimpan DataFrame ke Google Sheets (jika dikonfigurasi) dan/atau file Excel lokal.
    Mendukung penambahan kolom baru secara dinamis (skema fleksibel).
    """
    default_cols = ["tahun", "nama_opd", "penanggungjawab", "kode_rekening", "jenis_belanja", "bulan", "pagu_anggaran", "realisasi"]
    # Kolom turunan/hitung otomatis yang tidak perlu disimpan ke DB
    skip_cols = {"nama_bulan", "triwulan", "nama_triwulan", "sisa_anggaran", "persentase_realisasi", "pagu_fmt", "realisasi_fmt", "_select"}

    # Utamakan kolom default dulu, lalu tambahkan kolom kustom baru jika ada
    existing_cols = [c for c in default_cols if c in df.columns]
    extra_cols = [c for c in df.columns if c not in default_cols and c not in skip_cols and not c.startswith("_")]
    existing_cols.extend(extra_cols)

    clean_df = df[existing_cols].copy()

    if is_gsheets_configured():
        try:
            ws = get_gsheets_client_and_sheet()
            ws.clear()
            header = existing_cols
            values = [header] + clean_df.astype(str).values.tolist()
            ws.update("A1", values)
            st.cache_data.clear()
        except Exception as e:
            st.error(f"❌ Gagal menyimpan ke Google Sheets: {e}")

    # Simpan juga ke file Excel lokal sebagai cadangan
    path = filepath or DEFAULT_FILE
    os.makedirs(os.path.dirname(path), exist_ok=True)
    clean_df.to_excel(path, index=False, sheet_name="Realisasi Anggaran")


def add_record(
    tahun: int,
    jenis_belanja: str,
    bulan: int,
    pagu_anggaran: float,
    realisasi: float,
    penanggungjawab: str = "Sekretariat",
    kode_rekening: str = "",
    filepath: Optional[str] = None,
) -> pd.DataFrame:
    """
    Menambah satu baris data baru dengan validasi duplikasi.
    """
    df = load_raw_data(filepath)

    if not df.empty and "jenis_belanja" in df.columns:
        mask = (
            (pd.to_numeric(df["tahun"], errors="coerce") == tahun) &
            (pd.to_numeric(df["bulan"], errors="coerce") == bulan) &
            (df["jenis_belanja"].astype(str).str.strip().str.lower() == jenis_belanja.strip().lower())
        )
        if "kode_rekening" in df.columns and kode_rekening.strip():
            mask = mask & (df["kode_rekening"].astype(str).str.strip() == kode_rekening.strip())

        if mask.any():
            nama_bln = NAMA_BULAN.get(bulan, str(bulan))
            raise ValueError(
                f"Data realisasi untuk '{jenis_belanja}' pada bulan {nama_bln} {tahun} sudah pernah diinput. "
                f"Silakan gunakan menu '✏️ Edit Data' jika ingin memperbarui nilainya."
            )

    new_row = pd.DataFrame([{
        "tahun": tahun,
        "nama_opd": NAMA_OPD,
        "penanggungjawab": penanggungjawab,
        "kode_rekening": kode_rekening,
        "jenis_belanja": jenis_belanja,
        "bulan": bulan,
        "pagu_anggaran": pagu_anggaran,
        "realisasi": realisasi,
    }])

    df = pd.concat([df, new_row], ignore_index=True)
    df = df.sort_values(["tahun", "jenis_belanja", "bulan"]).reset_index(drop=True)
    save_data(df, filepath)
    return df


def update_record(
    idx: int,
    data: Dict[str, Any],
    filepath: Optional[str] = None,
) -> pd.DataFrame:
    """
    Mengupdate satu baris data berdasarkan index.
    """
    df = load_raw_data(filepath)

    if idx < 0 or idx >= len(df):
        raise ValueError(f"Index {idx} tidak valid.")

    for key, value in data.items():
        if key in df.columns:
            df.at[idx, key] = value

    save_data(df, filepath)
    return df


def delete_records(
    indices: List[int],
    filepath: Optional[str] = None,
) -> pd.DataFrame:
    """
    Menghapus baris data berdasarkan list index.
    """
    df = load_raw_data(filepath)
    df = df.drop(index=indices).reset_index(drop=True)
    save_data(df, filepath)
    return df


def bulk_save(df: pd.DataFrame, filepath: Optional[str] = None) -> None:
    """
    Menyimpan seluruh DataFrame yang sudah diedit (dari st.data_editor).
    """
    path = filepath or DEFAULT_FILE

    # Pastikan kolom nama_opd terisi
    if "nama_opd" not in df.columns:
        df["nama_opd"] = NAMA_OPD
    else:
        df["nama_opd"] = df["nama_opd"].fillna(NAMA_OPD)

    save_data(df, path)


def validate_and_sanitize_excel(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
    """
    Melakukan validasi dan sanitasi data Excel yang diupload sesuai best practice:
    1. Memastikan kolom utama tersedia.
    2. Membersihkan spasi pada string (jenis_belanja, penanggungjawab, kode_rekening).
    3. Mengonversi tipe data numerik (tahun, bulan, pagu_anggaran, realisasi).
    4. Mengembalikan DataFrame yang bersih dan list peringatan.
    """
    warnings = []
    clean_df = df.copy()

    clean_df.columns = clean_df.columns.astype(str).str.strip().str.lower().str.replace(" ", "_")

    required = ["tahun", "jenis_belanja", "bulan", "pagu_anggaran", "realisasi"]
    missing = [c for c in required if c not in clean_df.columns]
    if missing:
        return clean_df, [f"Kolom wajib tidak ditemukan: {', '.join(missing)}"]

    if "nama_opd" not in clean_df.columns:
        clean_df["nama_opd"] = NAMA_OPD
    else:
        clean_df["nama_opd"] = clean_df["nama_opd"].fillna(NAMA_OPD)

    if "penanggungjawab" not in clean_df.columns:
        clean_df["penanggungjawab"] = "Sekretariat"
    else:
        clean_df["penanggungjawab"] = clean_df["penanggungjawab"].fillna("Sekretariat").astype(str).str.strip()

    if "kode_rekening" not in clean_df.columns:
        clean_df["kode_rekening"] = ""
    else:
        clean_df["kode_rekening"] = clean_df["kode_rekening"].fillna("").astype(str).str.strip()

    clean_df["jenis_belanja"] = clean_df["jenis_belanja"].fillna("").astype(str).str.strip()

    clean_df["tahun"] = pd.to_numeric(clean_df["tahun"], errors="coerce").fillna(2025).astype(int)
    clean_df["bulan"] = pd.to_numeric(clean_df["bulan"], errors="coerce").fillna(1).astype(int)
    clean_df["pagu_anggaran"] = pd.to_numeric(clean_df["pagu_anggaran"], errors="coerce").fillna(0.0).abs()
    clean_df["realisasi"] = pd.to_numeric(clean_df["realisasi"], errors="coerce").fillna(0.0).abs()

    invalid_bulan = clean_df[(clean_df["bulan"] < 1) | (clean_df["bulan"] > 12)]
    if not invalid_bulan.empty:
        warnings.append(f"Terdapat {len(invalid_bulan)} baris dengan nilai bulan tidak valid (di luar 1-12).")
        clean_df["bulan"] = clean_df["bulan"].clip(1, 12)

    invalid_jenis = clean_df[clean_df["jenis_belanja"] == ""]
    if not invalid_jenis.empty:
        warnings.append(f"Terdapat {len(invalid_jenis)} baris dengan nama Sub-Kegiatan kosong.")

    return clean_df, warnings


def merge_save_records(new_df: pd.DataFrame, filepath: Optional[str] = None) -> int:
    """
    Menggabungkan (upsert) data Excel baru ke dalam database:
    - Jika kunci (tahun + bulan + jenis_belanja + kode_rekening) sudah ada, perbarui nilainya.
    - Jika belum ada, tambahkan sebagai baris baru.
    """
    path = filepath or DEFAULT_FILE
    existing_df = load_raw_data(path)

    if "nama_opd" not in new_df.columns:
        new_df["nama_opd"] = NAMA_OPD
    else:
        new_df["nama_opd"] = new_df["nama_opd"].fillna(NAMA_OPD)

    if existing_df.empty:
        bulk_save(new_df, path)
        return len(new_df)

    key_cols = ["tahun", "bulan", "jenis_belanja"]
    if "kode_rekening" in existing_df.columns and "kode_rekening" in new_df.columns:
        key_cols.append("kode_rekening")

    # new_df ditaruh di depan agar keep='first' mengambil nilai dari Excel baru
    combined = pd.concat([new_df, existing_df], ignore_index=True)
    combined = combined.drop_duplicates(subset=key_cols, keep="first")
    combined = combined.sort_values(["tahun", "jenis_belanja", "bulan"]).reset_index(drop=True)

    save_data(combined, path)
    return len(new_df)


def generate_formatted_excel_report(df_filtered: pd.DataFrame, summary: dict, tahun: int) -> bytes:
    """
    Menghasilkan file Excel terformat resmi laporan realisasi anggaran BKAD.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Laporan BKAD {tahun}"
    ws.views.sheetView[0].showGridLines = True

    HEADER_FILL = PatternFill(start_color="1B2838", end_color="1B2838", fill_type="solid")
    HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1B2838")
    SUBTITLE_FONT = Font(name="Arial", size=10, italic=True, color="555555")
    BOLD_FONT = Font(name="Arial", size=10, bold=True)
    REGULAR_FONT = Font(name="Arial", size=10)

    THIN_BORDER = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # 1. Title Block
    ws.merge_cells("A1:H1")
    ws["A1"] = "BADAN KEUANGAN DAN ASET DAERAH (BKAD)"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = f"LAPORAN REALISASI ANGGARAN TAHUN {tahun}"
    ws["A2"].font = Font(name="Arial", size=12, bold=True, color="2C3E50")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    latest_bln_name = NAMA_BULAN.get(summary.get("latest_bulan", 1), "")
    ws.merge_cells("A3:H3")
    ws["A3"] = f"Posisi Data s.d. Bulan {latest_bln_name} {tahun}"
    ws["A3"].font = SUBTITLE_FONT
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

    # 2. Summary Block
    ws["A5"] = "Ringkasan Eksekutif Realisasi:"
    ws["A5"].font = BOLD_FONT

    summary_rows = [
        ("Total Pagu Tahunan", summary.get("total_pagu", 0.0), "Rp #,##0"),
        ("Total Realisasi Kumulatif", summary.get("total_realisasi", 0.0), "Rp #,##0"),
        ("Total Sisa Anggaran", summary.get("total_sisa", 0.0), "Rp #,##0"),
        ("Persentase Capaian Overall", summary.get("persentase", 0.0) / 100.0, "0.00%"),
    ]

    r = 6
    for label, val, num_fmt in summary_rows:
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        cell_val = ws.cell(row=r, column=2, value=val)
        cell_val.font = BOLD_FONT
        cell_val.number_format = num_fmt
        r += 1

    # 3. Data Table Header
    r += 1
    headers = [
        "No", "Kode Rekening", "Penanggung Jawab", "Sub-Kegiatan / Uraian",
        "Pagu Anggaran (Rp)", "Realisasi Kumulatif (Rp)", "Sisa Anggaran (Rp)",
        "% Capaian"
    ]

    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 4. Populate Data Rows
    group_cols = ["jenis_belanja"]
    if "penanggungjawab" in df_filtered.columns:
        group_cols.insert(0, "penanggungjawab")
    if "kode_rekening" in df_filtered.columns:
        group_cols.insert(0, "kode_rekening")

    idx = df_filtered.groupby(group_cols, dropna=False)["bulan"].idxmax()
    latest_detail = df_filtered.loc[idx].copy()
    real_col = "realisasi_kumulatif" if "realisasi_kumulatif" in latest_detail.columns else "realisasi"
    latest_detail = latest_detail.sort_values("pagu_anggaran", ascending=False)

    for row_num, (_, row) in enumerate(latest_detail.iterrows(), 1):
        r += 1
        kode = str(row.get("kode_rekening", "")).replace(".0", "")
        pj = str(row.get("penanggungjawab", "Sekretariat"))
        jenis = str(row.get("jenis_belanja", ""))
        pagu = float(row.get("pagu_anggaran", 0.0))
        real = float(row.get(real_col, 0.0))
        sisa = pagu - real
        pct = (real / pagu * 100) if pagu > 0 else 0.0

        ws.cell(row=r, column=1, value=row_num).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=kode).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value=pj)
        ws.cell(row=r, column=4, value=jenis)

        c_pagu = ws.cell(row=r, column=5, value=pagu)
        c_pagu.number_format = "Rp #,##0"

        c_real = ws.cell(row=r, column=6, value=real)
        c_real.number_format = "Rp #,##0"

        c_sisa = ws.cell(row=r, column=7, value=sisa)
        c_sisa.number_format = "Rp #,##0"

        c_pct = ws.cell(row=r, column=8, value=pct / 100.0)
        c_pct.number_format = "0.00%"

        for col_i in range(1, 9):
            cell_item = ws.cell(row=r, column=col_i)
            cell_item.border = THIN_BORDER
            cell_item.font = REGULAR_FONT

    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_formatted_pdf_report(df_filtered: pd.DataFrame, summary: dict, tahun: int) -> bytes:
    """
    Menghasilkan laporan resmi Realisasi Anggaran BKAD berformat PDF.
    """
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from utils.charts import format_rupiah_titik

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=16,
        leading=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1B2838'),
    )

    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=11,
        leading=14,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#2C3E50'),
    )

    meta_style = ParagraphStyle(
        'DocMeta',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=9,
        leading=12,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#555555'),
    )

    h2_style = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=11,
        leading=14,
        textColor=colors.HexColor('#1B2838'),
        spaceAfter=6,
    )

    cell_style = ParagraphStyle(
        'TableCell',
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#2C3E50'),
    )

    cell_bold = ParagraphStyle(
        'CellBold',
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#1B2838'),
    )

    header_cell = ParagraphStyle(
        'HeaderCell',
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10,
        alignment=TA_CENTER,
        textColor=colors.white,
    )

    elements = []

    # Title Block
    elements.append(Paragraph("BADAN KEUANGAN DAN ASET DAERAH (BKAD)", title_style))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(f"LAPORAN REALISASI ANGGARAN TAHUN {tahun}", subtitle_style))
    elements.append(Spacer(1, 2))

    latest_bln_name = NAMA_BULAN.get(summary.get("latest_bulan", 1), "")
    elements.append(Paragraph(f"Posisi Data s.d. Bulan {latest_bln_name} {tahun}", meta_style))
    elements.append(Spacer(1, 8))
    elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#00E676'), spaceBefore=2, spaceAfter=10))

    # Executive Summary Card Table
    pagu_str = f"Rp {format_rupiah_titik(summary.get('total_pagu', 0.0))}"
    real_str = f"Rp {format_rupiah_titik(summary.get('total_realisasi', 0.0))}"
    sisa_str = f"Rp {format_rupiah_titik(summary.get('total_sisa', 0.0))}"
    pct_str = f"{summary.get('persentase', 0.0):.2f}%"

    sum_data = [
        [
            Paragraph("<b>Total Pagu Tahunan</b>", cell_bold),
            Paragraph(pagu_str, cell_bold),
            Paragraph("<b>Total Realisasi Kumulatif</b>", cell_bold),
            Paragraph(real_str, cell_bold),
        ],
        [
            Paragraph("<b>Total Sisa Anggaran</b>", cell_bold),
            Paragraph(sisa_str, cell_bold),
            Paragraph("<b>Persentase Capaian</b>", cell_bold),
            Paragraph(pct_str, cell_bold),
        ],
    ]

    sum_table = Table(sum_data, colWidths=[140, 180, 140, 180])
    sum_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F4F6F9')),
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('PADDING', (0,0), (-1,-1), 6),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elements.append(sum_table)
    elements.append(Spacer(1, 14))

    # Main Detail Table
    elements.append(Paragraph("Daftar Detail Realisasi Sub-Kegiatan BKAD:", h2_style))

    table_data = [
        [
            Paragraph("No", header_cell),
            Paragraph("Kode Rekening", header_cell),
            Paragraph("Penanggung Jawab", header_cell),
            Paragraph("Sub-Kegiatan / Uraian", header_cell),
            Paragraph("Pagu Anggaran", header_cell),
            Paragraph("Realisasi Kumulatif", header_cell),
            Paragraph("Sisa Anggaran", header_cell),
            Paragraph("% Capaian", header_cell),
        ]
    ]

    group_cols = ["jenis_belanja"]
    if "penanggungjawab" in df_filtered.columns:
        group_cols.insert(0, "penanggungjawab")
    if "kode_rekening" in df_filtered.columns:
        group_cols.insert(0, "kode_rekening")

    idx = df_filtered.groupby(group_cols, dropna=False)["bulan"].idxmax()
    latest_detail = df_filtered.loc[idx].copy()
    real_col = "realisasi_kumulatif" if "realisasi_kumulatif" in latest_detail.columns else "realisasi"
    latest_detail = latest_detail.sort_values("pagu_anggaran", ascending=False)

    right_cell = ParagraphStyle('RCell', parent=cell_style, alignment=TA_RIGHT)
    center_cell = ParagraphStyle('CCell', parent=cell_style, alignment=TA_CENTER)

    for row_num, (_, row) in enumerate(latest_detail.iterrows(), 1):
        kode = str(row.get("kode_rekening", "")).replace(".0", "")
        pj = str(row.get("penanggungjawab", "Sekretariat"))
        jenis = str(row.get("jenis_belanja", ""))
        pagu = float(row.get("pagu_anggaran", 0.0))
        real = float(row.get(real_col, 0.0))
        sisa = pagu - real
        pct = (real / pagu * 100) if pagu > 0 else 0.0

        table_data.append([
            Paragraph(str(row_num), center_cell),
            Paragraph(kode, center_cell),
            Paragraph(pj, cell_style),
            Paragraph(jenis, cell_style),
            Paragraph(f"Rp {format_rupiah_titik(pagu)}", right_cell),
            Paragraph(f"Rp {format_rupiah_titik(real)}", right_cell),
            Paragraph(f"Rp {format_rupiah_titik(sisa)}", right_cell),
            Paragraph(f"{pct:.2f}%", center_cell),
        ])

    pdf_table = Table(table_data, colWidths=[24, 75, 110, 240, 95, 95, 95, 45], repeatRows=1)
    pdf_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B2838')),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('PADDING', (0,0), (-1,-1), 4),
    ]))

    elements.append(pdf_table)

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
