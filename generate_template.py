"""
Script untuk membuat template Excel upload data realisasi anggaran BKAD.
Jalankan: python generate_template.py
"""
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Data contoh real (3 baris dari SIPD BKAD)
sample_data = [
    {
        "tahun": 2025,
        "nama_opd": "Badan Keuangan dan Aset Daerah (BKAD)",
        "penanggungjawab": "Sekretariat",
        "kode_rekening": "5.02.01.2.01.0001",
        "jenis_belanja": "Penyusunan Dokumen Perencanaan Perangkat Daerah",
        "bulan": 1,
        "pagu_anggaran": 16342500,
        "realisasi": 0,
    },
    {
        "tahun": 2025,
        "nama_opd": "Badan Keuangan dan Aset Daerah (BKAD)",
        "penanggungjawab": "Sekretariat",
        "kode_rekening": "5.02.01.2.01.0001",
        "jenis_belanja": "Penyusunan Dokumen Perencanaan Perangkat Daerah",
        "bulan": 2,
        "pagu_anggaran": 16342500,
        "realisasi": 2440000,
    },
    {
        "tahun": 2025,
        "nama_opd": "Badan Keuangan dan Aset Daerah (BKAD)",
        "penanggungjawab": "Bidang Anggaran",
        "kode_rekening": "5.02.02.2.01.0001",
        "jenis_belanja": "Koordinasi dan Penyusunan KUA dan PPAS",
        "bulan": 1,
        "pagu_anggaran": 325936500,
        "realisasi": 0,
    },
]

df = pd.DataFrame(sample_data)

os.makedirs("data", exist_ok=True)
filepath = "data/template_upload.xlsx"

# Simpan ke Excel - Sheet Data Realisasi sebagai Sheet 0 (Default Utama)
df.to_excel(filepath, index=False, sheet_name="Data Realisasi")

# ─── Styling dengan openpyxl ─────────────────────────────────────────────────

wb = load_workbook(filepath)
ws = wb["Data Realisasi"]

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

data_font = Font(name="Calibri", size=10)
data_alignment = Alignment(horizontal="left", vertical="center")
number_alignment = Alignment(horizontal="right", vertical="center")

thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

column_widths = {
    "A": 10,   # tahun
    "B": 40,   # nama_opd
    "C": 30,   # penanggungjawab
    "D": 20,   # kode_rekening
    "E": 45,   # jenis_belanja / uraian
    "F": 10,   # bulan
    "G": 20,   # pagu_anggaran
    "H": 20,   # realisasi
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Style headers
for col_idx in range(1, 9):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Style data rows (sample)
for row_idx in range(2, len(sample_data) + 2):
    for col_idx in range(1, 9):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = data_font
        cell.border = thin_border
        cell.fill = sample_fill

        if col_idx in [1, 6]:  # tahun, bulan
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx in [7, 8]:  # pagu, realisasi
            cell.alignment = number_alignment
            cell.number_format = '#,##0'
        else:
            cell.alignment = data_alignment

# Freeze header row
ws.freeze_panes = "A2"

# ─── Sheet Petunjuk (Sheet Index 1 / Setelah Data Realisasi) ────────────────

ws_info = wb.create_sheet("Petunjuk Pengisian", 1)

instructions = [
    ["PETUNJUK PENGISIAN TEMPLATE DATA REALISASI ANGGARAN BKAD"],
    [""],
    ["No", "Kolom", "Tipe Data", "Keterangan", "Contoh"],
    ["1", "tahun", "Angka (integer)", "Tahun anggaran (4 digit)", "2025"],
    ["2", "nama_opd", "Teks", "Nama lengkap OPD", "Badan Keuangan dan Aset Daerah (BKAD)"],
    ["3", "penanggungjawab", "Teks", "Bidang penanggung jawab kegiatan", "Sekretariat / Bidang Anggaran / Bidang Perbendaharaan"],
    ["4", "kode_rekening", "Teks", "Kode Rekening SIPD/SIMDA", "5.02.01.2.01.0001"],
    ["5", "jenis_belanja", "Teks", "Nama Sub-Kegiatan / Jenis Belanja", "Penyusunan Dokumen Perencanaan Perangkat Daerah"],
    ["6", "bulan", "Angka (1-12)", "Bulan realisasi (1 = Jan, 12 = Des)", "1"],
    ["7", "pagu_anggaran", "Angka (Rupiah)", "Pagu anggaran Rupiah penuh (tanpa titik/koma)", "16342500"],
    ["8", "realisasi", "Angka (Rupiah)", "Nilai realisasi bulanan mentah Rupiah penuh pada bulan berkenaan", "2440000"],
    [""],
    ["PENTING:"],
    ["1. Jangan mengubah nama kolom (baris pertama) pada sheet 'Data Realisasi'"],
    ["2. Data realisasi diisi per bulan (mentah/non-kumulatif). Sistem Dashboard akan menghitung akumulasi secara otomatis"],
    ["3. Pagu anggaran bernilai SAMA untuk semua bulan dalam 1 tahun per kegiatan"],
    ["4. Hapus 3 baris contoh (baris abu-abu) sebelum mengisi data asli"],
    ["5. Simpan file dalam format .xlsx"],
    [""],
    ["DAFTAR BIDANG PENANGGUNG JAWAB:"],
    ["- Sekretariat"],
    ["- Bidang Anggaran"],
    ["- Bidang Perbendaharaan"],
    ["- Bidang Akuntansi & Pelaporan"],
    ["- Bidang Pengelolaan BMD"],
]

for row_idx, row_data in enumerate(instructions, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

# Style judul
ws_info.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color="1B3A5C")

# Style header tabel
for col_idx in range(1, 6):
    cell = ws_info.cell(row=3, column=col_idx)
    cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

for row_idx in range(4, 12):
    for col_idx in range(1, 6):
        cell = ws_info.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

ws_info.cell(row=13, column=1).font = Font(name="Calibri", size=11, bold=True, color="E74C3C")

ws_info.column_dimensions["A"].width = 8
ws_info.column_dimensions["B"].width = 20
ws_info.column_dimensions["C"].width = 22
ws_info.column_dimensions["D"].width = 55
ws_info.column_dimensions["E"].width = 40

wb.active = 0  # Sheet 0 (Data Realisasi) aktif secara default
wb.save(filepath)

print(f"✅ Template SIPD BKAD berhasil dibuat: {filepath}")
